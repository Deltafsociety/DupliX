import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import os

class RoyalCompareApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Royal Compare")
        self.root.geometry("800x750") # Increased size for better layout to accommodate log
        self.root.resizable(True, True) # Allow resizing

        # Configure styles for a more modern look
        self.style = ttk.Style()
        self.style.theme_use('clam') # 'clam', 'alt', 'default', 'classic'
        self.style.configure('TFrame', background='#f0f0f0')
        self.style.configure('TLabel', background='#f0f0f0', font=('Inter', 10))
        self.style.configure('TButton', font=('Inter', 10, 'bold'), padding=8, background='#4CAF50', foreground='white')
        self.style.map('TButton', background=[('active', '#45a049')])
        self.style.configure('TEntry', font=('Inter', 10), padding=5)
        self.style.configure('TText', font=('Inter', 10))

        # --- Variables to store file paths and data ---
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()
        self.output_folder_path = tk.StringVar()
        self.df1_data = {} # Dictionary to hold DataFrames for each sheet of file 1
        self.df2_data = {} # Dictionary to hold DataFrames for each sheet of file 2
        self.common_numbers = set() # Set to store common numbers found

        self._create_widgets()

    def _create_widgets(self):
        # --- Main Frame ---
        main_frame = ttk.Frame(self.root, padding="20 20 20 20", relief="groove", borderwidth=2)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # --- File Selection Section ---
        file_selection_frame = ttk.Frame(main_frame, padding="10")
        file_selection_frame.grid(row=0, column=0, columnspan=2, pady=10, sticky="ew")
        file_selection_frame.columnconfigure(1, weight=1) # Make entry field expand

        # Updated label to clarify which file is highlighted
        ttk.Label(file_selection_frame, text="Select Excel File 1 (This will be the highlighted output):").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(file_selection_frame, textvariable=self.file1_path, width=50, state='readonly').grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(file_selection_frame, text="Browse...", command=lambda: self._browse_file(self.file1_path)).grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(file_selection_frame, text="Select Excel File 2:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(file_selection_frame, textvariable=self.file2_path, width=50, state='readonly').grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(file_selection_frame, text="Browse...", command=lambda: self._browse_file(self.file2_path)).grid(row=1, column=2, padx=5, pady=5)

        # --- Compare Button ---
        ttk.Button(main_frame, text="Compare and Highlight", command=self._compare_files, style='TButton').grid(row=1, column=0, columnspan=2, pady=15)

        # --- Output Folder Selection ---
        output_frame = ttk.Frame(main_frame, padding="10")
        output_frame.grid(row=2, column=0, columnspan=2, pady=10, sticky="ew")
        output_frame.columnconfigure(1, weight=1)

        ttk.Label(output_frame, text="Output Folder:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(output_frame, textvariable=self.output_folder_path, width=50, state='readonly').grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(output_frame, text="Browse...", command=self._browse_output_folder).grid(row=0, column=2, padx=5, pady=5)

        # --- Save Button ---
        ttk.Button(main_frame, text="Save Highlighted File 1", command=self._save_highlighted_files, style='TButton').grid(row=3, column=0, columnspan=2, pady=15)

        # --- Number Search Section ---
        search_frame = ttk.LabelFrame(main_frame, text="Search for a Number", padding="15")
        search_frame.grid(row=4, column=0, columnspan=2, pady=20, sticky="ew")
        search_frame.columnconfigure(1, weight=1)

        ttk.Label(search_frame, text="Enter Number:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.search_entry = ttk.Entry(search_frame, width=30)
        self.search_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ttk.Button(search_frame, text="Search", command=self._search_number, style='TButton').grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(search_frame, text="Search Results:").grid(row=1, column=0, padx=5, pady=5, sticky="nw")
        self.search_results_text = tk.Text(search_frame, height=5, width=60, wrap=tk.WORD, state='disabled')
        self.search_results_text.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        # --- Highlighting Process Log Section ---
        log_frame = ttk.LabelFrame(main_frame, text="Highlighting Process Log", padding="15")
        log_frame.grid(row=5, column=0, columnspan=2, pady=20, sticky="ew")
        log_frame.columnconfigure(0, weight=1)

        self.highlight_log_text = tk.Text(log_frame, height=7, width=70, wrap=tk.WORD, state='disabled')
        self.highlight_log_text.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        log_scrollbar = ttk.Scrollbar(log_frame, command=self.highlight_log_text.yview)
        log_scrollbar.grid(row=0, column=1, sticky='ns')
        self.highlight_log_text.config(yscrollcommand=log_scrollbar.set)


        # --- Status Bar ---
        self.status_label = ttk.Label(self.root, text="Ready", relief=tk.SUNKEN, anchor=tk.W, font=('Inter', 9))
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)

    def _browse_file(self, path_var):
        """Opens a file dialog to select an Excel file."""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            path_var.set(file_path)
            self.status_label.config(text=f"Selected: {os.path.basename(file_path)}")

    def _browse_output_folder(self):
        """Opens a directory dialog to select an output folder."""
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_folder_path.set(folder_path)
            self.status_label.config(text=f"Output folder set: {os.path.basename(folder_path)}")

    def _load_excel_data(self, file_path):
        """Loads all sheets from an Excel file into a dictionary of DataFrames."""
        if not file_path:
            return None
        try:
            # Use pd.ExcelFile to get sheet names and then read each sheet
            xls = pd.ExcelFile(file_path)
            data = {sheet_name: xls.parse(sheet_name) for sheet_name in xls.sheet_names}
            return data
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file '{file_path}': {e}")
            return None

    def _compare_files(self):
        """Compares the two selected Excel files and finds common numbers."""
        file1 = self.file1_path.get()
        file2 = self.file2_path.get()

        if not file1 or not file2:
            messagebox.showwarning("Warning", "Please select both Excel files.")
            return

        self.status_label.config(text="Comparing files...")
        self.root.update_idletasks() # Update GUI to show status

        self.df1_data = self._load_excel_data(file1)
        self.df2_data = self._load_excel_data(file2)

        if not self.df1_data or not self.df2_data:
            self.status_label.config(text="Comparison failed.")
            return

        all_numbers_file1 = set()
        all_numbers_file2 = set()

        # Extract all numeric values from File 1
        for sheet_name, df in self.df1_data.items():
            for col in df.columns:
                # Convert column to numeric, coercing errors to NaN
                numeric_series = pd.to_numeric(df[col], errors='coerce')
                # Drop NaN values and add to set
                all_numbers_file1.update(numeric_series.dropna().tolist())

        # Extract all numeric values from File 2
        for sheet_name, df in self.df2_data.items():
            for col in df.columns:
                numeric_series = pd.to_numeric(df[col], errors='coerce')
                all_numbers_file2.update(numeric_series.dropna().tolist())

        self.common_numbers = all_numbers_file1.intersection(all_numbers_file2)

        if self.common_numbers:
            self.status_label.config(text=f"Comparison complete. Found {len(self.common_numbers)} common numbers. Ready to save.")
            messagebox.showinfo("Success", f"Found {len(self.common_numbers)} common numbers. You can now save the highlighted files.")
        else:
            self.status_label.config(text="No common numbers found.")
            messagebox.showinfo("Result", "No common numbers found between the two files.")

    def _highlight_and_save_single_file(self, original_file_path, common_nums, output_folder, file_identifier, log_widget):
        """
        Loads an Excel file, highlights common numbers, and saves it to the output folder.
        Returns the path of the saved file or None on error.
        Logs highlighted numbers to the provided log_widget.
        """
        if not original_file_path or not os.path.exists(original_file_path):
            return None

        highlighted_count = 0
        file_base_name = os.path.basename(original_file_path)
        self._update_log(log_widget, f"--- Highlighting for {file_identifier} ({file_base_name}) ---")

        try:
            workbook = load_workbook(original_file_path)
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        # Check if the cell value is numeric and in common_nums
                        if isinstance(cell.value, (int, float)) and cell.value in common_nums:
                            cell.fill = red_fill
                            self._update_log(log_widget, f"Highlighted: {cell.value} in Sheet: '{sheet_name}'")
                            highlighted_count += 1

            # Generate new filename with timestamp
            name, ext = os.path.splitext(file_base_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"{name}_highlighted_{timestamp}{ext}"
            saved_file_path = os.path.join(output_folder, new_filename)

            workbook.save(saved_file_path)
            self._update_log(log_widget, f"Total numbers highlighted in {file_identifier}: {highlighted_count}")
            self._update_log(log_widget, f"Saved {file_identifier} to: {saved_file_path}\n")
            return saved_file_path

        except Exception as e:
            messagebox.showerror("Error", f"Failed to highlight and save '{file_base_name}': {e}")
            self._update_log(log_widget, f"Error highlighting {file_identifier}: {e}\n")
            return None

    def _save_highlighted_files(self):
        """Triggers the highlighting and saving process for the first file only."""
        if not self.common_numbers:
            messagebox.showwarning("Warning", "No common numbers found yet. Please run 'Compare and Highlight' first.")
            return

        output_folder = self.output_folder_path.get()
        if not output_folder or not os.path.isdir(output_folder):
            messagebox.showwarning("Warning", "Please select a valid output folder.")
            return

        self.status_label.config(text="Saving highlighted file 1...")
        self.root.update_idletasks()
        self._update_log(self.highlight_log_text, "Starting highlighting process...")

        # Only highlight and save the first file
        saved_path1 = self._highlight_and_save_single_file(
            self.file1_path.get(), self.common_numbers, output_folder, "File 1", self.highlight_log_text
        )

        if saved_path1:
            self.status_label.config(text="File 1 saved successfully!")
            messagebox.showinfo("Success", f"Highlighted File 1 saved to:\n{saved_path1}")
        else:
            self.status_label.config(text="Failed to save File 1.")
            messagebox.showerror("Error", "Failed to save the highlighted File 1. Check the log for details.")

    def _search_number(self):
        """Searches for a number in the loaded Excel files."""
        search_term = self.search_entry.get().strip()
        if not search_term:
            self._update_search_results("Please enter a number to search.")
            return

        try:
            num_to_find = float(search_term) # Allow for float numbers
        except ValueError:
            self._update_search_results("Invalid number. Please enter a numeric value.")
            return

        if not self.df1_data and not self.df2_data:
            self._update_search_results("No Excel files loaded. Please compare files first.")
            return

        results = []

        # Search in File 1
        if self.df1_data:
            file1_name = os.path.basename(self.file1_path.get())
            for sheet_name, df in self.df1_data.items():
                # Check if the number exists in any column of the current sheet
                # Convert all values to numeric for comparison, handling non-numeric gracefully
                if any((pd.to_numeric(df[col], errors='coerce') == num_to_find).any() for col in df.columns):
                    results.append(f"Found in File 1 ({file1_name}), Sheet: '{sheet_name}'")

        # Search in File 2
        if self.df2_data:
            file2_name = os.path.basename(self.file2_path.get())
            for sheet_name, df in self.df2_data.items():
                if any((pd.to_numeric(df[col], errors='coerce') == num_to_find).any() for col in df.columns):
                    results.append(f"Found in File 2 ({file2_name}), Sheet: '{sheet_name}'")

        if results:
            self._update_search_results("\n".join(results))
        else:
            self._update_search_results(f"Number '{search_term}' not found in any loaded files or sheets.")

    def _update_search_results(self, text):
        """Helper to update the search results text widget."""
        self.search_results_text.config(state='normal')
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, text)
        self.search_results_text.config(state='disabled')
        self.search_results_text.see(tk.END) # Scroll to the end

    def _update_log(self, log_widget, text):
        """Helper to update the log text widget."""
        log_widget.config(state='normal')
        log_widget.insert(tk.END, text + "\n")
        log_widget.config(state='disabled')
        log_widget.see(tk.END) # Scroll to the end

# --- Main execution ---
if __name__ == "__main__":
    root = tk.Tk()
    app = RoyalCompareApp(root)
    root.mainloop()
